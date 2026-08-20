"""Bulk youth intake from a spreadsheet — spec §4.1.

Woreda registers arrive as .xlsx, so registering a cohort one form at a time is
the wrong shape of work. This module turns a workbook into the same writes the
single-record form makes.

Three rules shape everything here:

* **Every row goes through `YouthIntakeSerializer`.** The consent rule (§9), the
  location vocabulary check and the age-band warning are not restated — a second
  copy of them would drift from the form's copy, and consent is the one thing
  that must not. This module's own job is only the spreadsheet: which cell is
  which field, and what an Excel value means.
* **A file is validated whole before anything is written.** `run_import` with
  ``commit=False`` is the preview the UI shows; the commit refuses outright if
  any row is invalid, inside one transaction. A half-imported register leaves
  nobody able to say which half.
* **Rows that already exist are skipped, not refused.** Registers get re-sent
  with twenty more names appended, and re-importing must not double the
  registry. That makes duplicates a normal outcome rather than an error.
"""

from __future__ import annotations

import datetime as dt
import re
from dataclasses import dataclass, field
from io import BytesIO

from django.db import transaction
from django.utils.translation import gettext_lazy as _

from .models import DisabilityStatus, EducationLevel, PsnpStatus, Sex, Youth
from .serializers import YouthIntakeSerializer

# The pilot is 500-1,000 youth in total (spec §1), so a single file larger than
# this is a mistake — a wrong sheet, or a register for the whole region. Refusing
# it is kinder than spending four minutes validating it.
MAX_ROWS = 2000
MAX_FILE_BYTES = 5 * 1024 * 1024

XLSX_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

TRUTHY = {"true", "t", "yes", "y", "1", "✓", "x"}
FALSEY = {"false", "f", "no", "n", "0", ""}


@dataclass(frozen=True)
class ImportColumn:
    """One spreadsheet column, and the Youth field it fills."""

    field: str
    header: str
    required: bool = False
    choices: type | None = None
    kind: str = "text"  # text | date | bool | choice
    note: str = ""

    @property
    def allowed(self) -> list[str]:
        return [value for value, _label in self.choices.choices] if self.choices else []


# Order is the order of the template's columns, so a user reading the sheet
# left to right meets the required ones first.
COLUMNS: list[ImportColumn] = [
    ImportColumn("full_name", "Full name", required=True, note="As written on the kebele register."),
    ImportColumn("sex", "Sex", required=True, kind="choice", choices=Sex),
    ImportColumn("date_of_birth", "Date of birth", required=True, kind="date", note="YYYY-MM-DD, or an Excel date."),
    ImportColumn("region", "Region", required=True, note="Must match the location reference data exactly."),
    ImportColumn("zone", "Zone", required=True, note="Must sit under the region."),
    ImportColumn("woreda", "Woreda", required=True, note="Must sit under the zone."),
    ImportColumn("kebele", "Kebele", required=True),
    ImportColumn(
        "consent_given",
        "Consent given",
        required=True,
        kind="bool",
        note="YES or NO. A youth cannot be registered without recorded consent (spec §9).",
    ),
    ImportColumn(
        "consent_date",
        "Consent date",
        required=True,
        kind="date",
        note="The date the youth gave consent — not the date of this import.",
    ),
    ImportColumn("phone_number", "Phone", note="Youth or next-of-kin contact."),
    ImportColumn("national_or_kebele_id", "National or kebele ID", note="Used to recognise a youth already imported."),
    ImportColumn("household_id", "PSNP household ID"),
    ImportColumn("psnp_status", "PSNP status", kind="choice", choices=PsnpStatus),
    ImportColumn("education_level", "Education level", kind="choice", choices=EducationLevel),
    ImportColumn("disability_status", "Disability status", kind="choice", choices=DisabilityStatus),
]

BY_HEADER = {column.header.casefold(): column for column in COLUMNS}
REQUIRED_HEADERS = [column.header for column in COLUMNS if column.required]

# Row outcomes. `new` and `duplicate` both mean "the file is fine"; only `error`
# blocks the commit.
NEW = "new"
DUPLICATE = "duplicate"
ERROR = "error"


class WorkbookError(Exception):
    """The file could not be read as a register at all."""


@dataclass
class RowResult:
    """One spreadsheet row, after validation."""

    number: int  # 1-based sheet row, so it matches what the user sees in Excel
    status: str
    full_name: str = ""
    errors: dict[str, list[str]] = field(default_factory=dict)
    warning: str = ""
    duplicate_of: str = ""  # id of the youth this row already exists as
    _payload: dict = field(default_factory=dict, repr=False)

    def as_dict(self) -> dict:
        return {
            "row": self.number,
            "status": self.status,
            "full_name": self.full_name,
            "errors": {key: [str(message) for message in messages] for key, messages in self.errors.items()},
            "warning": self.warning,
            "duplicate_of": self.duplicate_of or None,
        }


# ---------------------------------------------------------------------------
# Reading cells
# ---------------------------------------------------------------------------


def _clean(value) -> str:
    """A cell as trimmed text.

    Excel stores an ID typed as digits as a float, so `0912345678` comes back as
    `912345678.0` — the trailing `.0` has to go before it reaches a CharField, or
    every phone number in the file is subtly wrong.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (dt.datetime, dt.date)):
        return value.date().isoformat() if isinstance(value, dt.datetime) else value.isoformat()
    return str(value).strip()


def _coerce_date(value):
    """Return an ISO date string, or raise ValueError.

    openpyxl gives a `datetime` when the cell is date-formatted and a string when
    it is not, and registers are typed both ways in the same file.
    """
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()

    text = _clean(value)
    if not text:
        return ""
    # Accept the separators people actually type. Excel's own display format is
    # locale-dependent, which is exactly why the template asks for YYYY-MM-DD.
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"'{text}' is not a date. Use YYYY-MM-DD.")


def _coerce_bool(value):
    if isinstance(value, bool):
        return value
    text = _clean(value).casefold()
    if text in TRUTHY:
        return True
    if text in FALSEY:
        return False
    raise ValueError(f"'{text}' is not YES or NO.")


def _coerce_choice(value, column: ImportColumn):
    """Accept either the stored code or the English label, case-insensitively.

    Staff filling the register read the label; a file exported from this system
    carries the code. Both have to work, or a round trip through Excel fails.
    """
    text = _clean(value)
    if not text:
        return ""
    folded = text.casefold()
    for code, label in column.choices.choices:
        if folded in {code.casefold(), str(label).casefold()}:
            return code
    raise ValueError(f"'{text}' is not one of: {', '.join(column.allowed)}.")


def coerce_row(raw: dict, columns: list[ImportColumn]) -> tuple[dict, dict]:
    """`(payload, cell_errors)` — raw cells turned into Python values.

    `read_rows` deliberately returns raw cells, because the youth-side importer
    reports per-cell errors against the sheet row before anything is written.
    Any other caller has to do this step, and doing it by hand is how the WLT
    extract first arrived with `None` in every blank cell and ISO strings where
    dates belonged.
    """
    payload: dict = {}
    cell_errors: dict[str, list[str]] = {}
    for column in columns:
        if column.field not in raw:
            continue
        try:
            payload[column.field] = _cell_to_payload(column, raw[column.field])
        except ValueError as exc:
            cell_errors[column.field] = [str(exc)]
    return payload, cell_errors


def _cell_to_payload(column: ImportColumn, value):
    if column.kind == "date":
        return _coerce_date(value)
    if column.kind == "bool":
        return _coerce_bool(value)
    if column.kind == "choice":
        return _coerce_choice(value, column)
    return _clean(value)


def read_rows(stream, columns: list[ImportColumn] | None = None) -> list[tuple[int, dict]]:
    """Parse a workbook into `(sheet_row_number, {field: raw_cell})`.

    Reads the first worksheet only: the template ships one sheet of data plus a
    notes sheet, and picking "the first" is the rule a user can predict.

    `columns` defaults to the woreda register's. It is a parameter because the
    WLT module's PSNP ELS extract is a different sheet with the same problems —
    Excel's type confusion, headers in either case, unrecognised local columns —
    and a second copy of this would drift from it. What varies between the two
    registers is the column list; nothing else here does.
    """
    columns = columns or COLUMNS
    by_header = {column.header.casefold(): column for column in columns}
    required_headers = [column.header for column in columns if column.required]
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException

    try:
        # read_only keeps a 2,000-row file off the heap; data_only takes the
        # cached result of a formula rather than the formula text.
        workbook = load_workbook(stream, read_only=True, data_only=True)
    except InvalidFileException as exc:
        raise WorkbookError(_("That file is not a .xlsx workbook. Save it as Excel Workbook and try again.")) from exc
    except Exception as exc:  # openpyxl raises a wide spread on a corrupt zip
        raise WorkbookError(_("The file could not be opened as a spreadsheet.")) from exc

    try:
        sheet = workbook.worksheets[0]
        rows = sheet.iter_rows(values_only=True)

        try:
            header_row = next(rows)
        except StopIteration:
            raise WorkbookError(_("The first sheet is empty.")) from None

        # Map each column position to a field, ignoring columns we do not know.
        # An unrecognised column is not an error: registers carry local notes.
        positions: dict[int, ImportColumn] = {}
        for index, header in enumerate(header_row):
            column = by_header.get(_clean(header).casefold())
            if column:
                positions[index] = column

        missing = [header for header in required_headers if by_header[header.casefold()] not in positions.values()]
        if missing:
            raise WorkbookError(
                _("The sheet is missing these required columns: %(columns)s. Download the template.")
                % {"columns": ", ".join(missing)}
            )

        parsed: list[tuple[int, dict]] = []
        for offset, values in enumerate(rows, start=2):
            if all(_clean(value) == "" for value in values):
                continue  # a blank spacer row, not a youth
            parsed.append((offset, {column.field: values[index] for index, column in positions.items()}))
            if len(parsed) > MAX_ROWS:
                raise WorkbookError(
                    _("The file holds more than %(limit)s rows. Split it into smaller registers.") % {"limit": MAX_ROWS}
                )
        return parsed
    finally:
        workbook.close()


# ---------------------------------------------------------------------------
# Duplicate keys
# ---------------------------------------------------------------------------


def _norm(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip().casefold()


def _keys(payload: dict) -> list[tuple]:
    """The identities a row can already be known by.

    The ID is definitive where a register carries one. Where it does not, name
    plus date of birth is the pair a woreda office would themselves use to say
    "that is the same person" — weaker, but it stops the common re-send.
    """
    keys = []
    identifier = _norm(payload.get("national_or_kebele_id", ""))
    if identifier:
        keys.append(("id", identifier))
    name, born = _norm(payload.get("full_name", "")), payload.get("date_of_birth", "")
    if name and born:
        keys.append(("name_dob", name, str(born)))
    return keys


def _existing_index() -> dict[tuple, str]:
    """Every youth already on file, keyed the same way.

    Built unscoped and in one query: this is a duplicate check, not a read. A
    youth outside the importer's woreda is still a youth who must not be
    registered twice, and the row's own woreda still has to pass `_check_scope`.
    """
    index: dict[tuple, str] = {}
    for pk, name, born, identifier in Youth.objects.values_list(
        "pk", "full_name", "date_of_birth", "national_or_kebele_id"
    ):
        payload = {
            "full_name": name,
            "date_of_birth": born.isoformat() if born else "",
            "national_or_kebele_id": identifier,
        }
        for key in _keys(payload):
            index.setdefault(key, str(pk))
    return index


# ---------------------------------------------------------------------------
# The import itself
# ---------------------------------------------------------------------------


def _check_scope(payload: dict, user) -> list[str]:
    """Refuse a row that writes outside the importer's own woredas.

    Only applied to a user whose §7 case scope *is* a woreda list. A programme
    manager (Scope.ALL) imports anywhere; a case manager is scoped by caseload,
    which a youth who has no case yet cannot be measured against.

    TODO(open-question): `POST /youth/` does not make this check, so a single
    registration can still name any active woreda. Bulk is where it bites — one
    file can place a whole register in the wrong woreda's queue — but the two
    paths should agree. Not on the §11 list; raise it at the same Phase 1
    sign-off as the §7 scope gaps in CLAUDE.md's gotchas.
    """
    from apps.users.models import Scope

    if user.case_scope() != Scope.OWN_WOREDA:
        return []
    allowed = set(user.woreda_assignment or [])
    woreda = payload.get("woreda", "")
    if woreda and woreda not in allowed:
        return [
            _("'%(woreda)s' is outside your assigned woredas (%(allowed)s).")
            % {"woreda": woreda, "allowed": ", ".join(sorted(allowed)) or "none"}
        ]
    return []


def run_import(rows: list[tuple[int, dict]], user, *, request=None, commit: bool = False) -> dict:
    """Validate every row; write them only if all of them pass.

    Returns the same report either way, so the preview the user approves is
    literally the report the commit produces.
    """
    seen = _existing_index()
    results: list[RowResult] = []

    for number, raw in rows:
        result = RowResult(number=number, status=NEW)
        payload: dict = {}
        cell_errors: dict[str, list[str]] = {}

        for column in COLUMNS:
            if column.field not in raw:
                continue
            try:
                payload[column.field] = _cell_to_payload(column, raw[column.field])
            except ValueError as exc:
                cell_errors[column.field] = [str(exc)]

        result.full_name = str(payload.get("full_name", "") or "")

        # Blank optional cells must not reach the serializer as "": a CharField
        # accepts it, but a DateField and a choice field both reject it.
        payload = {key: value for key, value in payload.items() if value != "" or key == "full_name"}

        if cell_errors:
            result.status, result.errors = ERROR, cell_errors
            results.append(result)
            continue

        scope_errors = _check_scope(payload, user)
        if scope_errors:
            result.status, result.errors = ERROR, {"woreda": scope_errors}
            results.append(result)
            continue

        keys = _keys(payload)
        # Test for the key's presence, not its truth: a row claimed earlier in
        # this same file is stored with an empty id, and `if seen.get(key)`
        # would read that as "not seen" and write the youth twice.
        matched = next((key for key in keys if key in seen), None)
        if matched is not None:
            result.status, result.duplicate_of = DUPLICATE, seen[matched]
            results.append(result)
            continue

        serializer = YouthIntakeSerializer(data=payload, context={"request": request})
        if not serializer.is_valid():
            result.status, result.errors = ERROR, dict(serializer.errors)
            results.append(result)
            continue

        result._payload = serializer.validated_data
        # Claim the keys now, so a file that repeats a name inside itself reports
        # the second occurrence as a duplicate rather than writing it twice.
        for key in keys:
            seen.setdefault(key, "")
        results.append(result)

    counts = {
        "total": len(results),
        "new": sum(1 for row in results if row.status == NEW),
        "duplicate": sum(1 for row in results if row.status == DUPLICATE),
        "error": sum(1 for row in results if row.status == ERROR),
    }
    report = {"committed": False, "counts": counts, "rows": [row.as_dict() for row in results]}

    if not commit or counts["error"]:
        return report

    with transaction.atomic():
        for result in results:
            if result.status != NEW:
                continue
            youth = Youth(**result._payload, registering_worker=user)
            youth.save()
            # The age band is a warning, not a bar (§11) — carry it back so the
            # importer sees which of the rows they just wrote need confirming.
            if not youth.is_age_eligible:
                result.warning = str(
                    _("Age %(age)s is outside the youth band; eligibility needs confirmation.") % {"age": youth.age}
                )
    report["committed"] = True
    report["rows"] = [row.as_dict() for row in results]
    return report


# ---------------------------------------------------------------------------
# The template
# ---------------------------------------------------------------------------


def build_template(columns: list[ImportColumn] | None = None, title: str = "Youth") -> bytes:
    """The blank register, with its own instructions on a second sheet.

    Shipped from the same column list the parser reads, so the template cannot
    describe a column the importer does not accept. Parameterised for the same
    reason `read_rows` is — see its docstring.
    """
    columns = columns or COLUMNS
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title

    header_font = Font(bold=True, color="FFFFFF")
    # openpyxl takes ARGB hex only — no CSS custom property reaches a .xlsx.
    # This is the handoff's --green-700, the same fill the app's headers use.
    header_fill = PatternFill("solid", fgColor="1E5B3A")

    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=1, column=index, value=column.header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[cell.column_letter].width = max(14, min(30, len(column.header) + 8))

        if column.choices:
            # Quoted-list validation is capped at 255 characters by the format;
            # every list here is well inside it, but fall back rather than
            # writing a workbook Excel will call corrupt.
            formula = '"{}"'.format(",".join(column.allowed))
            if len(formula) <= 255:
                rule = DataValidation(type="list", formula1=formula, allow_blank=not column.required)
                sheet.add_data_validation(rule)
                rule.add(f"{cell.column_letter}2:{cell.column_letter}1001")

    sheet.freeze_panes = "A2"

    notes = workbook.create_sheet("Instructions")
    notes.column_dimensions["A"].width = 26
    notes.column_dimensions["B"].width = 12
    notes.column_dimensions["C"].width = 96

    intro = [
        ("How to use this register", "", ""),
        ("", "", "Fill one row per youth on the 'Youth' sheet. Do not rename or reorder the columns."),
        ("", "", "Consent must be recorded for every youth before the row can be imported (spec §9)."),
        ("", "", "Region, zone and woreda must match the platform's location list, and must nest correctly."),
        ("", "", "A youth already on file — same ID, or same name and date of birth — is skipped, not duplicated."),
        ("", "", f"At most {MAX_ROWS} rows per file."),
        ("", "", ""),
        ("Column", "Required", "Notes"),
    ]
    for row in intro:
        notes.append(row)
    notes["A1"].font = Font(bold=True, size=14)
    for cell in notes[len(intro)]:
        cell.font = Font(bold=True)

    for column in columns:
        note = column.note
        if column.choices:
            allowed = ", ".join(column.allowed)
            note = f"{note} One of: {allowed}." if note else f"One of: {allowed}."
        notes.append((column.header, "Yes" if column.required else "", note))

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
