"""Bulk youth intake from a spreadsheet — spec §4.1, consent from §9, scope from §7.

The behaviour worth pinning is not "openpyxl can read a file". It is that the
import cannot do anything the single-record form would refuse: no youth without
consent, no youth in a woreda the importer does not cover, no second copy of a
youth already on file, and nothing at all written from a file with a bad row.
"""

from datetime import date, timedelta
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from apps.youth.imports import COLUMNS, MAX_ROWS, build_template, read_rows, run_import
from apps.youth.models import Youth

pytestmark = pytest.mark.django_db

HEADERS = [column.header for column in COLUMNS]
IMPORT_URL = "/api/v1/youth/import/"
TEMPLATE_URL = "/api/v1/youth/import/template/"


def born(age):
    return (date.today() - timedelta(days=365 * age + 10)).isoformat()


def row(**overrides):
    """A valid register row, keyed by column header."""
    values = {
        "Full name": "Almaz Tesfaye",
        "Sex": "FEMALE",
        "Date of birth": born(22),
        "Region": "Oromia",
        "Zone": "East Shewa",
        "Woreda": "Adama",
        "Kebele": "Adama 01",
        "Consent given": "YES",
        "Consent date": date.today().isoformat(),
        "Phone": "0912345678",
        "National or kebele ID": "",
        "PSNP household ID": "",
        "PSNP status": "",
        "Education level": "",
        "Disability status": "",
    }
    values.update(overrides)
    return values


def workbook_bytes(rows, headers=HEADERS):
    book = Workbook()
    sheet = book.active
    sheet.append(list(headers))
    for values in rows:
        sheet.append([values.get(header, "") for header in headers])
    buffer = BytesIO()
    book.save(buffer)
    buffer.seek(0)
    return buffer


def upload(rows, headers=HEADERS):
    stream = workbook_bytes(rows, headers)
    stream.name = "register.xlsx"
    return stream


def parse(rows, headers=HEADERS):
    return read_rows(workbook_bytes(rows, headers))


# ---------------------------------------------------------------------------
# Reading the sheet
# ---------------------------------------------------------------------------


def test_template_headers_are_the_headers_the_parser_reads(locations, outreach_worker):
    """The template cannot describe a column the importer does not accept."""
    sheet = load_workbook(BytesIO(build_template())).worksheets[0]
    assert [cell.value for cell in sheet[1]] == HEADERS


def test_template_carries_its_own_instructions():
    book = load_workbook(BytesIO(build_template()))
    assert book.sheetnames == ["Youth", "Instructions"]
    notes = "\n".join(str(cell) for line in book["Instructions"].iter_rows(values_only=True) for cell in line if cell)
    assert "Consent" in notes
    # Every column is documented, or a user meets a rule only as a row error.
    for column in COLUMNS:
        assert column.header in notes


def test_a_missing_required_column_is_refused_whole(locations):
    from apps.youth.imports import WorkbookError

    without_consent = [header for header in HEADERS if header != "Consent given"]
    with pytest.raises(WorkbookError) as exc:
        parse([row()], headers=without_consent)
    assert "Consent given" in str(exc.value)


def test_unknown_columns_are_ignored(locations, outreach_worker):
    """Registers carry local notes; an extra column is not an error."""
    rows = parse([{**row(), "Kebele leader": "Bekele"}], headers=HEADERS + ["Kebele leader"])
    report = run_import(rows, outreach_worker, commit=True)
    assert report["counts"]["new"] == 1


def test_blank_rows_are_skipped(locations, outreach_worker):
    assert len(parse([row(), {}, row(**{"Full name": "Second Youth"})])) == 2


def test_a_row_number_points_at_the_sheet_row(locations, outreach_worker):
    """Errors are read next to Excel, so the number must be Excel's."""
    rows = parse([row(), row(**{"Sex": "Yes"})])
    report = run_import(rows, outreach_worker)
    assert [entry["row"] for entry in report["rows"]] == [2, 3]


# ---------------------------------------------------------------------------
# Cell values
# ---------------------------------------------------------------------------


def test_choices_accept_either_the_code_or_the_label(locations, outreach_worker):
    rows = parse([row(**{"Sex": "Female"}), row(**{"Full name": "B", "Sex": "MALE"})])
    report = run_import(rows, outreach_worker, commit=True)
    assert report["counts"]["error"] == 0
    assert set(Youth.objects.values_list("sex", flat=True)) == {"FEMALE", "MALE"}


def test_an_unknown_choice_names_what_is_allowed(locations, outreach_worker):
    report = run_import(parse([row(**{"Sex": "Lady"})]), outreach_worker)
    assert "FEMALE" in report["rows"][0]["errors"]["sex"][0]


def test_a_date_typed_as_text_is_accepted(locations, outreach_worker):
    report = run_import(parse([row(**{"Date of birth": "15/03/2003"})]), outreach_worker, commit=True)
    assert report["counts"]["new"] == 1
    assert Youth.objects.get().date_of_birth == date(2003, 3, 15)


def test_a_real_excel_date_cell_is_accepted(locations, outreach_worker):
    """openpyxl hands back a datetime when the cell is date-formatted."""
    import datetime as dt

    report = run_import(parse([row(**{"Date of birth": dt.datetime(2003, 3, 15)})]), outreach_worker, commit=True)
    assert Youth.objects.get().date_of_birth == date(2003, 3, 15)
    assert report["counts"]["new"] == 1


def test_a_phone_excel_stored_as_a_number_does_not_keep_its_decimal(locations, outreach_worker):
    run_import(parse([row(**{"Phone": 912345678.0})]), outreach_worker, commit=True)
    assert Youth.objects.get().phone_number == "912345678"


def test_an_unparseable_date_is_a_row_error_not_a_crash(locations, outreach_worker):
    report = run_import(parse([row(**{"Consent date": "last Tuesday"})]), outreach_worker)
    assert report["counts"]["error"] == 1
    assert "consent_date" in report["rows"][0]["errors"]


# ---------------------------------------------------------------------------
# The rules the form already enforces (§9 consent, §4.1 locations)
# ---------------------------------------------------------------------------


def test_a_youth_without_consent_is_refused(locations, outreach_worker):
    report = run_import(parse([row(**{"Consent given": "NO"})]), outreach_worker, commit=True)
    assert report["counts"]["error"] == 1
    assert "consent_given" in report["rows"][0]["errors"]
    assert not Youth.objects.exists()


def test_consent_without_a_date_is_refused(locations, outreach_worker):
    report = run_import(parse([row(**{"Consent date": ""})]), outreach_worker)
    assert report["rows"][0]["status"] == "error"


def test_an_unknown_woreda_is_refused(locations, programme_manager):
    report = run_import(parse([row(**{"Woreda": "Nowhere"})]), programme_manager)
    assert "woreda" in report["rows"][0]["errors"]


def test_a_woreda_under_the_wrong_zone_is_refused(locations, programme_manager):
    report = run_import(parse([row(**{"Zone": "Oromia"})]), programme_manager)
    assert report["rows"][0]["status"] == "error"


def test_the_importer_is_recorded_as_the_registering_worker(locations, outreach_worker):
    """§4.1 accountability: a client cannot choose whose name goes on the row."""
    run_import(parse([row()]), outreach_worker, commit=True)
    assert Youth.objects.get().registering_worker == outreach_worker


def test_an_out_of_band_age_is_written_but_flagged(locations, outreach_worker):
    """§11 leaves the band unconfirmed, so this warns rather than blocks."""
    report = run_import(parse([row(**{"Date of birth": born(41)})]), outreach_worker, commit=True)
    assert report["counts"]["new"] == 1
    assert "outside the youth band" in report["rows"][0]["warning"]


# ---------------------------------------------------------------------------
# All or nothing
# ---------------------------------------------------------------------------


def test_one_bad_row_writes_none_of_the_others(locations, outreach_worker):
    rows = parse([row(**{"Full name": "Good One"}), row(**{"Full name": "Bad One", "Consent given": "NO"})])
    report = run_import(rows, outreach_worker, commit=True)
    assert report["committed"] is False
    assert not Youth.objects.exists()


def test_a_dry_run_writes_nothing_and_reports_the_same_counts(locations, outreach_worker):
    rows = parse([row(), row(**{"Full name": "Second Youth"})])
    preview = run_import(rows, outreach_worker, commit=False)
    assert preview["committed"] is False and preview["counts"]["new"] == 2
    assert not Youth.objects.exists()

    committed = run_import(parse([row(), row(**{"Full name": "Second Youth"})]), outreach_worker, commit=True)
    assert committed["committed"] is True
    assert committed["counts"] == preview["counts"]
    assert Youth.objects.count() == 2


# ---------------------------------------------------------------------------
# Duplicates — a re-sent register must not double the registry
# ---------------------------------------------------------------------------


def test_a_youth_already_on_file_is_skipped_by_id(locations, outreach_worker, make_youth):
    existing = make_youth(name="Almaz Tesfaye", national_or_kebele_id="ID-77")
    report = run_import(parse([row(**{"National or kebele ID": "id-77 "})]), outreach_worker, commit=True)
    assert report["counts"] == {"total": 1, "new": 0, "duplicate": 1, "error": 0}
    assert report["rows"][0]["duplicate_of"] == str(existing.pk)
    assert Youth.objects.count() == 1


def test_a_youth_already_on_file_is_skipped_by_name_and_birthday(locations, outreach_worker, make_youth):
    existing = make_youth(name="Almaz Tesfaye", age=22)
    report = run_import(parse([row(**{"Date of birth": existing.date_of_birth.isoformat()})]), outreach_worker)
    assert report["rows"][0]["status"] == "duplicate"


def test_the_same_name_repeated_inside_one_file_is_written_once(locations, outreach_worker):
    report = run_import(parse([row(), row()]), outreach_worker, commit=True)
    assert report["counts"]["new"] == 1 and report["counts"]["duplicate"] == 1
    assert Youth.objects.count() == 1


def test_re_importing_the_same_register_changes_nothing(locations, outreach_worker):
    """The register gets re-sent with more names appended; that must be safe."""
    run_import(parse([row(), row(**{"Full name": "Second Youth"})]), outreach_worker, commit=True)
    again = run_import(
        parse([row(), row(**{"Full name": "Second Youth"}), row(**{"Full name": "Third Youth"})]),
        outreach_worker,
        commit=True,
    )
    assert again["counts"]["duplicate"] == 2 and again["counts"]["new"] == 1
    assert Youth.objects.count() == 3


def test_two_different_youth_sharing_a_name_are_kept_apart_by_birthday(locations, outreach_worker):
    rows = parse([row(), row(**{"Date of birth": born(25)})])
    report = run_import(rows, outreach_worker, commit=True)
    assert report["counts"]["new"] == 2


# ---------------------------------------------------------------------------
# Scope — §7
# ---------------------------------------------------------------------------


def test_an_outreach_worker_cannot_import_into_another_woreda(locations, outreach_worker):
    """One wrong file would otherwise drop a whole register into another woreda."""
    report = run_import(parse([row(**{"Woreda": "Bishoftu"})]), outreach_worker, commit=True)
    assert report["counts"]["error"] == 1
    assert "outside your assigned woredas" in report["rows"][0]["errors"]["woreda"][0]
    assert not Youth.objects.exists()


def test_a_user_scoped_to_everything_imports_any_woreda(locations, system_admin):
    """Scope.ALL has no woreda list to be outside of, so the check does not apply."""
    report = run_import(parse([row(**{"Woreda": "Bishoftu"})]), system_admin, commit=True)
    assert report["counts"]["new"] == 1


# ---------------------------------------------------------------------------
# The endpoints
# ---------------------------------------------------------------------------


def test_the_template_downloads_as_a_workbook(locations, outreach_worker, as_user):
    response = as_user(outreach_worker).get(TEMPLATE_URL)
    assert response.status_code == 200
    assert "attachment" in response["Content-Disposition"]
    assert load_workbook(BytesIO(response.content)).sheetnames == ["Youth", "Instructions"]


def test_upload_previews_without_writing(locations, outreach_worker, as_user):
    response = as_user(outreach_worker).post(IMPORT_URL, {"file": upload([row()])}, format="multipart")
    assert response.status_code == 200
    assert response.data["committed"] is False
    assert not Youth.objects.exists()


def test_upload_with_commit_writes(locations, outreach_worker, as_user):
    response = as_user(outreach_worker).post(f"{IMPORT_URL}?commit=true", {"file": upload([row()])}, format="multipart")
    assert response.data["committed"] is True
    assert Youth.objects.count() == 1


def test_upload_without_a_file_is_a_400(locations, outreach_worker, as_user):
    assert as_user(outreach_worker).post(IMPORT_URL, {}, format="multipart").status_code == 400


def test_a_file_that_is_not_a_workbook_is_a_400(locations, outreach_worker, as_user):
    junk = BytesIO(b"woreda register, comma separated")
    junk.name = "register.xlsx"
    response = as_user(outreach_worker).post(IMPORT_URL, {"file": junk}, format="multipart")
    assert response.status_code == 400
    assert "spreadsheet" in response.data["detail"] or "xlsx" in response.data["detail"]


def test_a_header_only_sheet_is_a_400(locations, outreach_worker, as_user):
    response = as_user(outreach_worker).post(IMPORT_URL, {"file": upload([])}, format="multipart")
    assert response.status_code == 400


def test_a_reader_cannot_import(locations, supervisor, as_user):
    """§7 gives the supervisor read on case records; import is a write."""
    response = as_user(supervisor).post(IMPORT_URL, {"file": upload([row()])}, format="multipart")
    assert response.status_code == 403
    assert not Youth.objects.exists()


def test_a_programme_manager_cannot_import(locations, programme_manager, as_user):
    """§7 gives Scope.ALL for reading; writing case content is still refused."""
    response = as_user(programme_manager).post(IMPORT_URL, {"file": upload([row()])}, format="multipart")
    assert response.status_code == 403
    assert not Youth.objects.exists()


def test_a_system_administrator_can_import(locations, system_admin, as_user):
    """Pins the ACCESS_MATRIX deviation rather than §7 as written.

    §7 makes the administrator configuration-only; the matrix was widened to full
    case write on 2026-08-16 at the programme's request, and carries a
    TODO(spec-deviation) to that effect. If that widening is reversed at Phase 1
    sign-off, this test is the one that should fail and be flipped.
    """
    response = as_user(system_admin).post(IMPORT_URL, {"file": upload([row()])}, format="multipart")
    assert response.status_code == 200


def test_import_needs_authentication(locations, api):
    assert api.post(IMPORT_URL, {"file": upload([row()])}, format="multipart").status_code == 401


def test_a_file_over_the_row_limit_is_refused(locations, outreach_worker):
    from apps.youth.imports import WorkbookError

    many = [row(**{"Full name": f"Youth {index}"}) for index in range(MAX_ROWS + 1)]
    with pytest.raises(WorkbookError) as exc:
        parse(many)
    assert str(MAX_ROWS) in str(exc.value)
