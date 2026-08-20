"""Both ways into the register, and the journey that reads them back.

Decision D5 describes hybrid enrolment: import the ELS caseload, and let a
facilitator add the women it missed. Both halves existed as services and neither
had a route — the import had no parser and no endpoint, and the exception route
needed a `youth.Youth` row that no WLT role can create. So the module could only
be populated from a shell.
"""

from datetime import date
from io import BytesIO

import pytest
from openpyxl import Workbook

from apps.users.models import User
from apps.wlt.imports import COLUMNS
from apps.wlt.models import BeneficiaryProfile, EnrolmentRoute, GroupMembership, VerificationStatus
from apps.wlt.services import formation as formation_service
from apps.youth.models import Sex, Youth

pytestmark = pytest.mark.django_db


REGISTRATION = {
    "full_name": "Almaz Tesfaye",
    "date_of_birth": "1988-04-11",
    "consent_given": True,
    "consent_date": "2026-02-01",
    "phone_number": "0911234567",
    "psnp_client_id": "PSNP-99001",
    "els_completed_on": "2025-11-01",
    "els_grant_received_on": "2025-12-01",
    "has_device": True,
}


def _register(client, kebele, **overrides):
    payload = {**REGISTRATION, "kebele": kebele.code, **overrides}
    return client.post("/api/v1/wlt/profiles/register/", payload, format="json")


# ---------------------------------------------------------------------------
# The facilitator's exception route
# ---------------------------------------------------------------------------


def test_a_facilitator_registers_a_woman_the_extract_missed(as_user, facilitator, wlt_group, wlt_locations):
    response = _register(as_user(facilitator), wlt_locations["kebele"])

    assert response.status_code == 201
    profile = BeneficiaryProfile.objects.get(person__full_name="Almaz Tesfaye")
    assert profile.enrolment_route == EnrolmentRoute.FACILITATOR
    # Rule 3: the exception route starts pending, or it becomes the main route.
    assert profile.verification_status == VerificationStatus.PENDING
    assert profile.psnp_kebele_id == wlt_locations["kebele"].pk


def test_registration_derives_the_place_from_the_kebele(as_user, facilitator, wlt_group, wlt_locations):
    """Never accepted from the caller. A hand-typed woreda that disagrees with
    its kebele scopes to one place and reports in another."""
    _register(as_user(facilitator), wlt_locations["kebele"], woreda="Somewhere Else", region="Tigray")

    person = Youth.objects.get(full_name="Almaz Tesfaye")
    assert person.woreda == wlt_locations["woreda"].name
    assert person.region == wlt_locations["region"].name
    assert person.sex == Sex.FEMALE


def test_a_registration_without_consent_is_refused(as_user, facilitator, wlt_group, wlt_locations):
    """§9 makes consent the basis for holding the record at all."""
    response = _register(as_user(facilitator), wlt_locations["kebele"], consent_given=False)

    assert response.status_code == 400
    assert not Youth.objects.filter(full_name="Almaz Tesfaye").exists()


def test_a_second_woman_with_the_same_psnp_id_is_refused(as_user, facilitator, wlt_group, wlt_locations):
    client = as_user(facilitator)
    assert _register(client, wlt_locations["kebele"]).status_code == 201

    again = _register(client, wlt_locations["kebele"], full_name="Almaz Tesfaye Two")

    assert again.status_code == 400
    assert "psnp_client_id" in again.data


def test_a_shared_name_is_not_refused(as_user, facilitator, wlt_group, wlt_locations):
    """Rule 2 forbids turning a name match into a decision.

    Two women in one kebele really can share a name, and refusing the second is
    as wrong as merging her. The duplicate check that matters is at group
    assignment, where one open membership per person is a database constraint.
    """
    client = as_user(facilitator)
    assert _register(client, wlt_locations["kebele"]).status_code == 201

    namesake = _register(client, wlt_locations["kebele"], psnp_client_id="PSNP-99002")

    assert namesake.status_code == 201
    assert Youth.objects.filter(full_name="Almaz Tesfaye").count() == 2


def test_a_kebele_is_required_not_a_woreda(as_user, facilitator, wlt_group, wlt_locations):
    response = _register(as_user(facilitator), wlt_locations["woreda"])

    assert response.status_code == 400
    assert "kebele" in response.data


def test_a_facilitator_cannot_register_outside_the_kebeles_she_works(
    as_user, other_facilitator, wlt_group, wlt_locations
):
    """Fails closed, like every other scope decision in the module."""
    response = _register(as_user(other_facilitator), wlt_locations["kebele"])

    assert response.status_code == 400
    assert not Youth.objects.filter(full_name="Almaz Tesfaye").exists()


def test_a_woreda_officer_registers_anywhere_under_her_woreda(as_user, woreda_officer, wlt_locations):
    assert _register(as_user(woreda_officer), wlt_locations["other_kebele"]).status_code == 201


def test_registering_does_not_open_a_case_door(as_user, facilitator, wlt_group, wlt_locations):
    """The boundary `test_boundary` pins is unchanged by this route.

    Writing a `Youth` row is not reading a case file, and the whole reason this
    endpoint exists rather than a widened `CanAccessCases` is that the two can
    be told apart.
    """
    client = as_user(facilitator)
    assert _register(client, wlt_locations["kebele"]).status_code == 201

    assert client.get("/api/v1/cases/").status_code == 403
    assert client.get("/api/v1/youth/").status_code == 403


def test_a_case_manager_cannot_register_into_the_wlt_register(as_user, case_manager, wlt_locations):
    assert _register(as_user(case_manager), wlt_locations["kebele"]).status_code == 403


def test_an_officer_may_enrol_without_being_able_to_write_group_records(
    as_user, region_officer, wlt_group, wlt_members, wlt_locations
):
    """`CanEnrolBeneficiaries` is narrower than it looks.

    An officer holds the extract and verifies against it, so she enrols. She
    still cannot touch a group record, a meeting or the ledger — `group_write`
    is false for her and every other route keeps reading it.
    """
    client = as_user(region_officer)

    assert _register(client, wlt_locations["kebele"]).status_code == 201

    membership = GroupMembership.objects.get(group=wlt_group, person=wlt_members[0])
    assert (
        client.post(f"/api/v1/wlt/groups/{wlt_group.pk}/members/{membership.pk}/exit/", {"reason": "MOVED"}).status_code
        == 403
    )
    assert client.post(f"/api/v1/wlt/groups/{wlt_group.pk}/activate/", {}).status_code == 403


# ---------------------------------------------------------------------------
# The ELS extract
# ---------------------------------------------------------------------------


def _extract(rows):
    """A workbook whose header row is the template's, in the template's order."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([column.header for column in COLUMNS])
    by_field = {column.field: index for index, column in enumerate(COLUMNS)}
    for row in rows:
        cells = [None] * len(COLUMNS)
        for field, value in row.items():
            cells[by_field[field]] = value
        sheet.append(cells)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    buffer.name = "extract.xlsx"
    return buffer


def test_an_extract_lands_and_its_women_arrive_verified(as_user, woreda_officer, wlt_locations):
    """An imported woman is verified on arrival; only the exception route pends."""
    upload = _extract(
        [
            {"full_name": "Bekelech Alemu", "date_of_birth": "1985-03-02", "els_completed_on": "2025-10-01"},
            {"full_name": "Kidist Haile", "date_of_birth": "1990-07-19", "els_completed_on": "2025-10-05"},
        ]
    )

    response = as_user(woreda_officer).post(
        "/api/v1/wlt/profiles/import/",
        {"file": upload, "kebele": wlt_locations["kebele"].code, "batch": "els-2026-q1"},
        format="multipart",
    )

    assert response.status_code == 200, response.data
    assert response.data["outcomes"]["created"] == 2, response.data
    profile = BeneficiaryProfile.objects.get(person__full_name="Bekelech Alemu")
    assert profile.enrolment_route == EnrolmentRoute.IMPORT
    assert profile.verification_status == VerificationStatus.VERIFIED


def test_the_same_extract_twice_creates_nothing_new(as_user, woreda_officer, wlt_locations):
    """Idempotent on the PSNP client id — extracts get re-sent."""
    rows = [
        {
            "full_name": "Bekelech Alemu",
            "date_of_birth": "1985-03-02",
            "els_completed_on": "2025-10-01",
            "psnp_client_id": "PSNP-4001",
        }
    ]
    client = as_user(woreda_officer)
    payload = {"kebele": wlt_locations["kebele"].code, "batch": "els-2026-q1"}

    first = client.post("/api/v1/wlt/profiles/import/", {"file": _extract(rows), **payload}, format="multipart")
    second = client.post("/api/v1/wlt/profiles/import/", {"file": _extract(rows), **payload}, format="multipart")

    assert first.data["outcomes"]["created"] == 1
    assert second.data["outcomes"]["skipped"] == 1
    assert BeneficiaryProfile.objects.filter(psnp_client_id="PSNP-4001").count() == 1


def test_a_sheet_missing_a_required_column_is_refused_with_the_column_named(as_user, woreda_officer, wlt_locations):
    workbook = Workbook()
    workbook.active.append(["Full name", "Phone"])
    workbook.active.append(["Bekelech Alemu", "0911000000"])
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    buffer.name = "extract.xlsx"

    response = as_user(woreda_officer).post(
        "/api/v1/wlt/profiles/import/",
        {"file": buffer, "kebele": wlt_locations["kebele"].code},
        format="multipart",
    )

    assert response.status_code == 400
    assert "Date of birth" in str(response.data)


def test_a_row_with_an_unreadable_cell_is_rejected_not_half_imported(as_user, woreda_officer, wlt_locations):
    """Dropping the bad cell and importing anyway would be worse than not importing.

    A woman registered with no birth date is a record nothing downstream ever
    calls incomplete. She is named against her sheet row instead, and the rest
    of the extract still lands.
    """
    upload = _extract(
        [
            {"full_name": "Good Row", "date_of_birth": "1985-03-02", "els_completed_on": "2025-10-01"},
            {"full_name": "Bad Date", "date_of_birth": "the third of never", "els_completed_on": "2025-10-01"},
        ]
    )

    response = as_user(woreda_officer).post(
        "/api/v1/wlt/profiles/import/",
        {"file": upload, "kebele": wlt_locations["kebele"].code},
        format="multipart",
    )

    assert response.status_code == 200
    assert response.data["outcomes"]["created"] == 1
    assert [row["row"] for row in response.data["unreadable"]] == [3]
    assert "date_of_birth" in response.data["unreadable"][0]["errors"]
    assert Youth.objects.filter(full_name="Good Row").exists()
    assert not Youth.objects.filter(full_name="Bad Date").exists()


def test_the_template_is_built_from_the_columns_the_parser_reads(as_user, woreda_officer):
    response = as_user(woreda_officer).get("/api/v1/wlt/profiles/import-template/")

    assert response.status_code == 200
    assert response["Content-Disposition"].endswith('filename="psnp-els-extract-template.xlsx"')

    from openpyxl import load_workbook

    sheet = load_workbook(BytesIO(response.content)).worksheets[0]
    headers = [cell.value for cell in sheet[1]]
    assert headers == [column.header for column in COLUMNS]


def test_an_extract_cannot_be_loaded_into_a_kebele_outside_scope(as_user, other_facilitator, wlt_locations):
    upload = _extract([{"full_name": "Bekelech Alemu", "date_of_birth": "1985-03-02"}])

    response = as_user(other_facilitator).post(
        "/api/v1/wlt/profiles/import/",
        {"file": upload, "kebele": wlt_locations["kebele"].code},
        format="multipart",
    )

    assert response.status_code == 400
    assert not Youth.objects.filter(full_name="Bekelech Alemu").exists()


# ---------------------------------------------------------------------------
# The journey
# ---------------------------------------------------------------------------


def _journey(client, profile):
    response = client.get(f"/api/v1/wlt/profiles/{profile.pk}/journey/")
    assert response.status_code == 200
    return {stage["code"]: stage for stage in response.data["stages"]}, response.data


def test_the_journey_walks_registered_verified_grouped_linked(as_user, facilitator, wlt_group, wlt_members):
    profile = BeneficiaryProfile.objects.get(person=wlt_members[0])

    stages, payload = _journey(as_user(facilitator), profile)

    assert [stage["code"] for stage in payload["stages"]] == ["REGISTERED", "VERIFIED", "GROUPED", "LINKED"]
    assert stages["REGISTERED"]["state"] == "done"
    assert stages["VERIFIED"]["state"] == "done"
    assert stages["GROUPED"]["state"] == "done"
    assert stages["GROUPED"]["detail"]["group_name"] == wlt_group.name


def test_a_pending_woman_is_waiting_not_blocked(as_user, facilitator, wlt_group, wlt_locations):
    """Verification is a woreda officer's decision.

    `waiting` and `blocked` call for different things from the facilitator most
    likely to be reading the screen, and only one of them is her move.
    """
    _register(as_user(facilitator), wlt_locations["kebele"])
    profile = BeneficiaryProfile.objects.get(person__full_name="Almaz Tesfaye")

    stages, payload = _journey(as_user(facilitator), profile)

    assert stages["VERIFIED"]["state"] == "waiting"
    # She cannot be seated while pending, and the group stage says which
    # condition is the problem rather than simply refusing later.
    assert stages["GROUPED"]["state"] == "blocked"
    unmet = [c["code"] for c in stages["GROUPED"]["conditions"] if not c["met"]]
    assert unmet == ["verified"]
    assert payload["next_action"]["code"] == "VERIFIED"


def test_the_group_stage_names_every_condition_add_member_would_refuse(
    as_user, facilitator, wlt_group, make_wlt_member
):
    """The whole point: the four eligibility conditions readable in advance."""
    ineligible = make_wlt_member("No Grant Yet", els_grant_received_on=None)
    profile = BeneficiaryProfile.objects.get(person=ineligible)

    stages, _payload = _journey(as_user(facilitator), profile)

    codes = [condition["code"] for condition in stages["GROUPED"]["conditions"]]
    assert codes == ["verified", "female", "els_completed", "els_grant", "psnp"]
    unmet = [c["code"] for c in stages["GROUPED"]["conditions"] if not c["met"]]
    assert unmet == ["els_grant"]
    assert stages["GROUPED"]["state"] == "blocked"


def test_an_eligible_ungrouped_woman_is_ready(as_user, facilitator, wlt_group, make_wlt_member):
    joiner = make_wlt_member("Ready To Join")
    profile = BeneficiaryProfile.objects.get(person=joiner)

    stages, payload = _journey(as_user(facilitator), profile)

    assert stages["GROUPED"]["state"] == "ready"
    assert payload["next_action"]["code"] == "GROUPED"


def test_the_linkage_stage_names_the_phase_each_type_needs(as_user, facilitator, wlt_group, wlt_members):
    """ "All the gates included": a type her group cannot reach yet says why.

    A facilitator asking why the bank option is absent gets "needs Phase 2,
    group is at Phase 1" rather than an empty list.
    """
    profile = BeneficiaryProfile.objects.get(person=wlt_members[0])

    stages, _payload = _journey(as_user(facilitator), profile)
    linked = stages["LINKED"]

    assert linked["detail"]["group_name"] == wlt_group.name
    named = linked["detail"]["available_types"] + linked["detail"]["blocked_types"]
    assert named, "every active linkage type permitting a group should be named"
    for row in linked["detail"]["blocked_types"]:
        assert row["min_phase"], "a type is only blocked here because of the phase it needs"


def test_a_woman_in_no_group_cannot_be_linked_and_says_so(as_user, facilitator, wlt_group, make_wlt_member):
    profile = BeneficiaryProfile.objects.get(person=make_wlt_member("Ungrouped"))

    stages, _payload = _journey(as_user(facilitator), profile)

    assert stages["LINKED"]["state"] == "blocked"
    assert [c["code"] for c in stages["LINKED"]["conditions"]] == ["in_group"]


def test_the_journey_reflects_an_exit_immediately(as_user, facilitator, wlt_group, wlt_members):
    """Computed on request, like the readiness card, for the same reason."""
    person = wlt_members[-1]
    profile = BeneficiaryProfile.objects.get(person=person)
    membership = GroupMembership.objects.get(group=wlt_group, person=person)

    before, _ = _journey(as_user(facilitator), profile)
    formation_service.exit_member(membership, reason="MOVED", on_date=date(2026, 4, 1))
    after, _ = _journey(as_user(facilitator), profile)

    assert before["GROUPED"]["state"] == "done"
    assert after["GROUPED"]["state"] == "ready"


def test_a_case_manager_cannot_read_a_womans_journey(as_user, case_manager, wlt_group, wlt_members):
    profile = BeneficiaryProfile.objects.get(person=wlt_members[0])

    assert as_user(case_manager).get(f"/api/v1/wlt/profiles/{profile.pk}/journey/").status_code == 403


def test_a_facilitator_with_no_scope_registers_nowhere(as_user, wlt_locations, db):
    """Fails closed. A facilitator with no group and no assignment has no place."""
    from apps.users.models import Role

    unassigned = User.objects.create_user(
        "wlt-new", "pw-Test-12345", full_name="New Facilitator", role=Role.WLT_FACILITATOR
    )

    assert _register(as_user(unassigned), wlt_locations["kebele"]).status_code == 400
